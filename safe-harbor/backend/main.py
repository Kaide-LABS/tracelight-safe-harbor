import os
import uuid
import json
import asyncio
import logging
from fastapi import FastAPI, UploadFile, File, Form, WebSocket, WebSocketDisconnect, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from backend.config import get_settings
from backend.orchestrator import PipelineOrchestrator
from backend.models.schemas import JobState
from backend.middleware.logging_middleware import StructuredLoggingMiddleware
from backend.health import router as health_router

from fastapi.staticfiles import StaticFiles

logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

app = FastAPI()
app.include_router(health_router)
app.add_middleware(StructuredLoggingMiddleware)

app.mount("/templates", StaticFiles(directory="templates"), name="templates")
settings = get_settings()

allowed_origins = [
    "http://localhost:5173",
    "http://localhost:5174",
    "http://localhost:5175",
    os.getenv("FRONTEND_ORIGIN", ""),
]
# Add Cloud Run origins
for env_key in ["LAUNCHER_URL", "SAFE_HARBOR_FRONTEND_URL"]:
    url = os.getenv(env_key, "")
    if url:
        allowed_origins.append(url)
# Allow all *.run.app origins for Cloud Run
allowed_origins.append("https://*.run.app")
allowed_origins = [o for o in allowed_origins if o]

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_origin_regex=r"https://.*\.run\.app",
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

orchestrator = PipelineOrchestrator(settings)


def _get_google_creds(sa_path: str):
    """Load OAuth user credentials (preferred) or fall back to service account."""
    import os
    token_path = os.path.join(os.path.dirname(sa_path), 'oauth_token.json')
    if os.path.exists(token_path):
        from google.oauth2.credentials import Credentials
        creds = Credentials.from_authorized_user_file(token_path)
        if creds and creds.expired and creds.refresh_token:
            from google.auth.transport.requests import Request
            creds.refresh(Request())
            # Save refreshed token
            import json
            with open(token_path, 'w') as f:
                json.dump({
                    'token': creds.token,
                    'refresh_token': creds.refresh_token,
                    'token_uri': creds.token_uri,
                    'client_id': creds.client_id,
                    'client_secret': creds.client_secret,
                    'scopes': list(creds.scopes or []),
                }, f)
        return creds
    else:
        from google.oauth2 import service_account
        SCOPES = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']
        return service_account.Credentials.from_service_account_file(sa_path, scopes=SCOPES)


def _create_sheet_from_xlsx(xlsx_path: str, title: str, sa_path: str, add_validation: bool = False, parsed_template: dict = None, conformance_report: dict = None) -> dict:
    """Read xlsx with openpyxl, create Google Sheet via Sheets API, write all data."""
    import openpyxl
    from googleapiclient.discovery import build

    creds = _get_google_creds(sa_path)
    sheets_svc = build('sheets', 'v4', credentials=creds)
    drive_svc = build('drive', 'v3', credentials=creds)

    # Read xlsx
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)

    # Create spreadsheet with correct sheet names
    sheet_props = [{"properties": {"title": ws.title}} for ws in wb.worksheets]
    body = {"properties": {"title": title}, "sheets": sheet_props}
    spreadsheet = sheets_svc.spreadsheets().create(body=body, fields='spreadsheetId').execute()
    spreadsheet_id = spreadsheet['spreadsheetId']

    # Write data sheet by sheet
    data = []
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            row_data = []
            for cell in row:
                val = cell.value
                if val is None:
                    row_data.append({"userEnteredValue": {"stringValue": ""}})
                elif isinstance(val, str) and val.startswith("="):
                    row_data.append({"userEnteredValue": {"formulaValue": val}})
                elif isinstance(val, bool):
                    row_data.append({"userEnteredValue": {"boolValue": val}})
                elif isinstance(val, (int, float)):
                    row_data.append({"userEnteredValue": {"numberValue": val}})
                else:
                    row_data.append({"userEnteredValue": {"stringValue": str(val)}})
            rows.append({"values": row_data})

        # Find the sheet ID
        sheet_meta = sheets_svc.spreadsheets().get(
            spreadsheetId=spreadsheet_id, fields='sheets.properties'
        ).execute()
        sheet_id = None
        for s in sheet_meta['sheets']:
            if s['properties']['title'] == ws.title:
                sheet_id = s['properties']['sheetId']
                break

        if sheet_id is not None:
            data.append({
                "updateCells": {
                    "rows": rows,
                    "fields": "userEnteredValue",
                    "start": {"sheetId": sheet_id, "rowIndex": 0, "columnIndex": 0},
                }
            })

    if data:
        sheets_svc.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": data},
        ).execute()

    # Add Validation sheet with live formulas (only for generated output, not templates)
    if add_validation:
        _add_validation_sheet(sheets_svc, spreadsheet_id, wb, parsed_template=parsed_template, conformance_report=conformance_report)

    # Make publicly viewable
    drive_svc.permissions().create(
        fileId=spreadsheet_id,
        body={'type': 'anyone', 'role': 'reader'},
    ).execute()

    embed_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit?embedded=true&rm=minimal"
    view_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}"

    return {"embed_url": embed_url, "view_url": view_url, "sheet_id": spreadsheet_id}


def _add_validation_sheet(sheets_svc, spreadsheet_id: str, wb, parsed_template: dict = None, conformance_report: dict = None):
    """Add a 'Validation' sheet with live formulas proving data integrity.
    Uses row_map from parsed_template for all row references — no hardcoded numbers."""
    import re
    from backend.agents.row_map import build_row_map

    # Build row map from parser data if available
    rm = build_row_map(parsed_template) if parsed_template else None

    # Detect sheets — use row_map if available, else fall back to name matching
    wb_sheet_names = [ws.title for ws in wb.worksheets]
    wb_names_lower = {s.lower(): s for s in wb_sheet_names}

    if rm:
        sn = rm["sheet_names"]
        is_name = sn.get("is")
        bs_name = sn.get("bs")
        cf_name = sn.get("cf")
        ds_name = sn.get("ds")
        ra_name = sn.get("ra")
        template_type = rm["template_type"]
        row_map = rm["row_map"]
    else:
        is_name = next((v for k, v in wb_names_lower.items() if "income" in k), None)
        bs_name = next((v for k, v in wb_names_lower.items() if "balance" in k), None)
        cf_name = next((v for k, v in wb_names_lower.items() if "cash flow" in k), None)
        ds_name = next((v for k, v in wb_names_lower.items() if "debt" in k), None)
        ra_name = next((v for k, v in wb_names_lower.items() if "return" in k), None)
        template_type = "unknown"
        row_map = {}

    has_is = is_name is not None
    has_bs = bs_name is not None
    has_cf = cf_name is not None
    has_ds = ds_name is not None

    def _r(sheet_name, canonical):
        """Get row number for a canonical key, or None."""
        if not sheet_name or not rm:
            return None
        return row_map.get((sheet_name, canonical))

    # Detect periods
    periods = []
    if has_is:
        ws = wb[is_name]
        for col in range(2, ws.max_column + 1):
            for r in [1, 2]:
                val = ws.cell(row=r, column=col).value
                if val and re.search(r'(FY|CY)?\d{4}', str(val)):
                    periods.append({"col_letter": chr(64 + col), "label": str(val).strip()})
                    break
    if not periods:
        return

    cols = [p["col_letter"] for p in periods]

    # Add the validation sheet
    sheets_svc.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"requests": [{"addSheet": {"properties": {"title": "\u2713 Validation"}}}]},
    ).execute()

    # Build validation rows
    rows = []
    current_row = 0  # Track row number for self-references

    def _header(text):
        return [{"userEnteredValue": {"stringValue": text}, "userEnteredFormat": {"textFormat": {"bold": True, "fontSize": 11}}}]

    def _label(text):
        return [{"userEnteredValue": {"stringValue": text}}]

    def _formula_row(label, formulas):
        r = [{"userEnteredValue": {"stringValue": label}}]
        for f in formulas:
            if f and f.startswith("="):
                r.append({"userEnteredValue": {"formulaValue": f}})
            else:
                r.append({"userEnteredValue": {"stringValue": str(f) if f else ""}})
        return r

    def _status_row(label, check_formulas):
        r = [{"userEnteredValue": {"stringValue": label}}]
        for f in check_formulas:
            r.append({"userEnteredValue": {"formulaValue": f}})
        return r

    def _add(row_values):
        nonlocal current_row
        rows.append({"values": row_values})
        current_row += 1

    # Title
    _add(_header("SAFE-HARBOR VALIDATION REPORT"))
    _add(_label("All checks are live Google Sheets formulas \u2014 click any cell to verify."))
    _add([])  # blank
    # Period headers row
    period_row = [{"userEnteredValue": {"stringValue": ""}}]
    for p in periods:
        period_row.append({"userEnteredValue": {"stringValue": p["label"]}, "userEnteredFormat": {"textFormat": {"bold": True}}})
    _add(period_row)

    # ── Section 1: Balance Sheet Identity (UNIVERSAL) ──
    _add([])
    _add(_header("1. BALANCE SHEET IDENTITY (Assets = Liabilities + Equity)"))
    r_ta = _r(bs_name, "bs_total_assets")
    r_tl = _r(bs_name, "bs_total_liab")
    r_te = _r(bs_name, "bs_total_equity")
    if has_bs and all([r_ta, r_tl, r_te]):
        bs = bs_name
        _add(_formula_row("Total Assets", [f"='{bs}'!{c}{r_ta}" for c in cols]))
        _add(_formula_row("Total Liabilities", [f"='{bs}'!{c}{r_tl}" for c in cols]))
        _add(_formula_row("Total Equity", [f"='{bs}'!{c}{r_te}" for c in cols]))
        _add(_formula_row("\u0394 (Assets - L - E)", [f"='{bs}'!{c}{r_ta}-('{bs}'!{c}{r_tl}+'{bs}'!{c}{r_te})" for c in cols]))
        _add(_status_row("\u2713 Status", [f"=IF(ABS('{bs}'!{c}{r_ta}-('{bs}'!{c}{r_tl}+'{bs}'!{c}{r_te}))<1,\"PASS\",\"FAIL\")" for c in cols]))

    # ── Section 2: Margin Analysis (UNIVERSAL) ──
    _add([])
    _add(_header("2. MARGIN ANALYSIS"))
    r_rev = _r(is_name, "is_revenue")
    r_gp = _r(is_name, "is_gross_profit")
    r_ebitda = _r(is_name, "is_ebitda")
    r_ni = _r(is_name, "is_net_income")
    if has_is and r_rev:
        isn = is_name
        if r_gp:
            _add(_formula_row("Revenue", [f"='{isn}'!{c}{r_rev}" for c in cols]))
            _add(_formula_row("Gross Profit", [f"='{isn}'!{c}{r_gp}" for c in cols]))
            _add(_formula_row("Gross Margin %", [f"=IFERROR('{isn}'!{c}{r_gp}/'{isn}'!{c}{r_rev},\"N/A\")" for c in cols]))
        if r_ebitda:
            _add(_formula_row("EBITDA", [f"='{isn}'!{c}{r_ebitda}" for c in cols]))
            _add(_formula_row("EBITDA Margin %", [f"=IFERROR('{isn}'!{c}{r_ebitda}/'{isn}'!{c}{r_rev},\"N/A\")" for c in cols]))
        if r_ni:
            _add(_formula_row("Net Income", [f"='{isn}'!{c}{r_ni}" for c in cols]))
            _add(_formula_row("Net Margin %", [f"=IFERROR('{isn}'!{c}{r_ni}/'{isn}'!{c}{r_rev},\"N/A\")" for c in cols]))
        if r_gp:
            _add(_status_row("\u2713 Margins in range", [f"=IF(AND(IFERROR('{isn}'!{c}{r_gp}/'{isn}'!{c}{r_rev},0)>0,IFERROR('{isn}'!{c}{r_gp}/'{isn}'!{c}{r_rev},0)<1),\"PASS\",\"FAIL\")" for c in cols]))

    # ── Section 3: Revenue Growth (UNIVERSAL) ──
    _add([])
    _add(_header("3. REVENUE GROWTH RATE"))
    if has_is and r_rev and len(cols) > 1:
        isn = is_name
        # Skip first period (no prior year), wrap in IFERROR
        growth_formulas = [""]  # blank for FY2020
        for i in range(1, len(cols)):
            growth_formulas.append(f"=IFERROR('{isn}'!{cols[i]}{r_rev}/'{isn}'!{cols[i-1]}{r_rev}-1,\"N/A\")")
        _add(_formula_row("YoY Growth %", growth_formulas))
        # Track the row number where growth was written for Avg Growth reference
        growth_row = current_row  # 1-indexed in Sheets (current_row already incremented)
        if len(cols) > 2:
            _add(_formula_row("Avg Growth", ["", f"=IFERROR(AVERAGE({cols[1]}{growth_row}:{cols[-1]}{growth_row}),\"N/A\")"]))
        else:
            _add(_formula_row("Avg Growth", [""]))

    # ── Section 4: Cash Flow Reconciliation (UNIVERSAL where CF exists) ──
    _add([])
    _add(_header("4. CASH FLOW RECONCILIATION"))
    r_cf_begin = _r(cf_name, "cf_begin_cash")
    r_cf_net = _r(cf_name, "cf_net_change")
    r_cf_end = _r(cf_name, "cf_end_cash")
    if has_cf and all([r_cf_begin, r_cf_net, r_cf_end]):
        cfn = cf_name
        _add(_formula_row("Beginning Cash", [f"='{cfn}'!{c}{r_cf_begin}" for c in cols]))
        _add(_formula_row("Net Change in Cash", [f"='{cfn}'!{c}{r_cf_net}" for c in cols]))
        _add(_formula_row("Ending Cash", [f"='{cfn}'!{c}{r_cf_end}" for c in cols]))
        _add(_formula_row("\u0394 (End - Begin - Net)", [f"='{cfn}'!{c}{r_cf_end}-'{cfn}'!{c}{r_cf_begin}-'{cfn}'!{c}{r_cf_net}" for c in cols]))
        _add(_status_row("\u2713 Status", [f"=IF(ABS('{cfn}'!{c}{r_cf_end}-'{cfn}'!{c}{r_cf_begin}-'{cfn}'!{c}{r_cf_net})<1,\"PASS\",\"FAIL\")" for c in cols]))

    # ── Section 5: Debt Schedule Rollforward (LBO ONLY) ──
    if template_type == "LBO" and has_ds:
        r_sb = _r(ds_name, "ds_senior_begin")
        r_sd = _r(ds_name, "ds_senior_draw")
        r_sr = _r(ds_name, "ds_senior_repay")
        r_se = _r(ds_name, "ds_senior_end")
        if all([r_sb, r_sd, r_sr, r_se]):
            dsn = ds_name
            _add([])
            _add(_header("5. DEBT SCHEDULE \u2014 SENIOR"))
            _add(_formula_row("Beginning Balance", [f"='{dsn}'!{c}{r_sb}" for c in cols]))
            _add(_formula_row("+ Drawdowns", [f"='{dsn}'!{c}{r_sd}" for c in cols]))
            _add(_formula_row("- Repayments", [f"='{dsn}'!{c}{r_sr}" for c in cols]))
            _add(_formula_row("= Ending Balance", [f"='{dsn}'!{c}{r_se}" for c in cols]))
            # FIXED: End - (Begin + Draw - Repay) — repayments REDUCE the balance
            _add(_formula_row("\u0394 Rollforward", [f"='{dsn}'!{c}{r_se}-('{dsn}'!{c}{r_sb}+'{dsn}'!{c}{r_sd}-'{dsn}'!{c}{r_sr})" for c in cols]))
            _add(_status_row("\u2713 Status", [f"=IF(ABS('{dsn}'!{c}{r_se}-('{dsn}'!{c}{r_sb}+'{dsn}'!{c}{r_sd}-'{dsn}'!{c}{r_sr}))<1,\"PASS\",\"FAIL\")" for c in cols]))

        # Mezz rollforward
        r_mb = _r(ds_name, "ds_mezz_begin")
        r_md = _r(ds_name, "ds_mezz_draw")
        r_mr = _r(ds_name, "ds_mezz_repay")
        r_me = _r(ds_name, "ds_mezz_end")
        if all([r_mb, r_md, r_mr, r_me]):
            dsn = ds_name
            _add([])
            _add(_header("5b. DEBT SCHEDULE \u2014 MEZZANINE"))
            _add(_formula_row("Beginning Balance", [f"='{dsn}'!{c}{r_mb}" for c in cols]))
            _add(_formula_row("+ Drawdowns", [f"='{dsn}'!{c}{r_md}" for c in cols]))
            _add(_formula_row("- Repayments", [f"='{dsn}'!{c}{r_mr}" for c in cols]))
            _add(_formula_row("= Ending Balance", [f"='{dsn}'!{c}{r_me}" for c in cols]))
            _add(_formula_row("\u0394 Rollforward", [f"='{dsn}'!{c}{r_me}-('{dsn}'!{c}{r_mb}+'{dsn}'!{c}{r_md}-'{dsn}'!{c}{r_mr})" for c in cols]))
            _add(_status_row("\u2713 Status", [f"=IF(ABS('{dsn}'!{c}{r_me}-('{dsn}'!{c}{r_mb}+'{dsn}'!{c}{r_md}-'{dsn}'!{c}{r_mr}))<1,\"PASS\",\"FAIL\")" for c in cols]))

    # ── Section 6: Cross-Sheet Linkage ──
    _add([])
    _add(_header("6. CROSS-SHEET LINKAGE"))
    r_is_da = _r(is_name, "is_da")
    r_cf_da = _r(cf_name, "cf_da")
    if has_is and has_cf and r_is_da and r_cf_da:
        isn = is_name
        cfn = cf_name
        _add(_formula_row("IS: D&A", [f"='{isn}'!{c}{r_is_da}" for c in cols]))
        _add(_formula_row("CF: D&A Add-back", [f"='{cfn}'!{c}{r_cf_da}" for c in cols]))
        _add(_formula_row("\u0394 (IS D&A - CF D&A)", [f"=ABS('{isn}'!{c}{r_is_da})-ABS('{cfn}'!{c}{r_cf_da})" for c in cols]))
        _add(_status_row("\u2713 D&A Linkage", [f"=IF(ABS(ABS('{isn}'!{c}{r_is_da})-ABS('{cfn}'!{c}{r_cf_da}))<1,\"PASS\",\"FAIL\")" for c in cols]))

    r_is_int = _r(is_name, "is_interest_expense")
    r_ds_int = _r(ds_name, "ds_total_interest")
    if has_is and has_ds and r_is_int and r_ds_int:
        isn = is_name
        dsn = ds_name
        _add(_formula_row("IS: Total Interest", [f"='{isn}'!{c}{r_is_int}" for c in cols]))
        _add(_formula_row("DS: Total Interest", [f"='{dsn}'!{c}{r_ds_int}" for c in cols]))
        _add(_status_row("\u2713 Interest Linkage", [f"=IF(ABS(ABS('{isn}'!{c}{r_is_int})-ABS('{dsn}'!{c}{r_ds_int}))<1,\"PASS\",\"FAIL\")" for c in cols]))

    # ── Section 7: Statistical Summary (UNIVERSAL) ──
    _add([])
    _add(_header("7. STATISTICAL DISTRIBUTION"))
    if has_is and r_rev:
        isn = is_name
        rev_range = f"'{isn}'!{cols[0]}{r_rev}:{cols[-1]}{r_rev}"
        _add([
            {"userEnteredValue": {"stringValue": "Revenue"}},
            {"userEnteredValue": {"stringValue": "Mean"}},
            {"userEnteredValue": {"formulaValue": f"=AVERAGE({rev_range})"}},
            {"userEnteredValue": {"stringValue": "Std Dev"}},
            {"userEnteredValue": {"formulaValue": f"=STDEV({rev_range})"}},
            {"userEnteredValue": {"stringValue": "CV"}},
            {"userEnteredValue": {"formulaValue": f"=IFERROR(STDEV({rev_range})/AVERAGE({rev_range}),\"N/A\")"}},
        ])
        # Gross Margin computed inline (no margin row in template)
        if r_gp:
            _add([
                {"userEnteredValue": {"stringValue": "Gross Margin"}},
                {"userEnteredValue": {"stringValue": "Mean"}},
                {"userEnteredValue": {"formulaValue": f"=IFERROR(AVERAGE(ARRAYFORMULA('{isn}'!{cols[0]}{r_gp}:{cols[-1]}{r_gp}/'{isn}'!{cols[0]}{r_rev}:{cols[-1]}{r_rev})),\"N/A\")"}},
                {"userEnteredValue": {"stringValue": "Min"}},
                {"userEnteredValue": {"formulaValue": f"=IFERROR(MIN(ARRAYFORMULA('{isn}'!{cols[0]}{r_gp}:{cols[-1]}{r_gp}/'{isn}'!{cols[0]}{r_rev}:{cols[-1]}{r_rev})),\"N/A\")"}},
                {"userEnteredValue": {"stringValue": "Max"}},
                {"userEnteredValue": {"formulaValue": f"=IFERROR(MAX(ARRAYFORMULA('{isn}'!{cols[0]}{r_gp}:{cols[-1]}{r_gp}/'{isn}'!{cols[0]}{r_rev}:{cols[-1]}{r_rev})),\"N/A\")"}},
            ])

    # ── Section 8: Leverage Ratios (LBO ONLY) ──
    if template_type == "LBO" and has_ds and has_is:
        r_se = _r(ds_name, "ds_senior_end")
        r_ebitda_is = _r(is_name, "is_ebitda")
        r_ebit_is = _r(is_name, "is_ebit")
        r_is_int2 = _r(is_name, "is_interest_expense")
        if r_se and r_ebitda_is:
            dsn = ds_name
            isn = is_name
            _add([])
            _add(_header("8. LEVERAGE & COVERAGE RATIOS"))
            _add(_formula_row("Total Debt (Senior End)", [f"='{dsn}'!{c}{r_se}" for c in cols]))
            _add(_formula_row("Debt / EBITDA", [f"=IFERROR('{dsn}'!{c}{r_se}/'{isn}'!{c}{r_ebitda_is},\"N/A\")" for c in cols]))
            if r_ebit_is and r_is_int2:
                _add(_formula_row("Interest Coverage (EBIT/Interest)", [f"=IFERROR('{isn}'!{c}{r_ebit_is}/ABS('{isn}'!{c}{r_is_int2}),\"N/A\")" for c in cols]))

    # ── Section 9: Archetype Conformance Report ──
    if conformance_report and conformance_report.get("metrics"):
        _add([])
        _add(_header(f"9. ARCHETYPE CONFORMANCE — {conformance_report.get('scenario_type', 'general').upper().replace('_', ' ')}"))
        _add(_label(f"Overall Score: {conformance_report.get('overall_score', 'N/A')} ({conformance_report.get('pass_rate_pct', 0)}%)"))
        _add([])
        # Header row
        conf_header = [
            {"userEnteredValue": {"stringValue": "Metric"}, "userEnteredFormat": {"textFormat": {"bold": True}}},
            {"userEnteredValue": {"stringValue": "Period"}, "userEnteredFormat": {"textFormat": {"bold": True}}},
            {"userEnteredValue": {"stringValue": "Expected Range"}, "userEnteredFormat": {"textFormat": {"bold": True}}},
            {"userEnteredValue": {"stringValue": "Actual"}, "userEnteredFormat": {"textFormat": {"bold": True}}},
            {"userEnteredValue": {"stringValue": "Status"}, "userEnteredFormat": {"textFormat": {"bold": True}}},
            {"userEnteredValue": {"stringValue": "Source"}, "userEnteredFormat": {"textFormat": {"bold": True}}},
        ]
        _add(conf_header)
        for m in conformance_report["metrics"]:
            expected = m.get("expected_range", "")
            if isinstance(expected, list) and len(expected) == 2:
                expected_str = f"{expected[0]} — {expected[1]}"
            else:
                expected_str = str(expected)
            actual_val = m.get("actual")
            actual_str = str(actual_val) if actual_val is not None else "N/A"
            status = m.get("status", "N/A")
            _add([
                {"userEnteredValue": {"stringValue": m.get("name", "")}},
                {"userEnteredValue": {"stringValue": m.get("period", "")}},
                {"userEnteredValue": {"stringValue": expected_str}},
                {"userEnteredValue": {"stringValue": actual_str}},
                {"userEnteredValue": {"stringValue": status}},
                {"userEnteredValue": {"stringValue": m.get("source", "")}},
            ])

    # Write the validation sheet
    sheet_meta = sheets_svc.spreadsheets().get(
        spreadsheetId=spreadsheet_id, fields='sheets.properties'
    ).execute()
    val_sheet_id = None
    for s in sheet_meta['sheets']:
        if s['properties']['title'] == '\u2713 Validation':
            val_sheet_id = s['properties']['sheetId']
            break

    if val_sheet_id is not None and rows:
        sheets_svc.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": [{
                "updateCells": {
                    "rows": rows,
                    "fields": "userEnteredValue,userEnteredFormat",
                    "start": {"sheetId": val_sheet_id, "rowIndex": 0, "columnIndex": 0},
                }
            }]},
        ).execute()


@app.on_event("startup")
async def startup():
    os.makedirs("/tmp/safe_harbor", exist_ok=True)

@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...), scenario_type: str = Form("general")):
    if not file.filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Must be an .xlsx or .xlsm file")

    valid_scenarios = ["general", "distressed_turnaround", "high_growth_tech", "mature_cashcow"]
    if scenario_type not in valid_scenarios:
        scenario_type = "general"

    job_id = str(uuid.uuid4())
    job_dir = f"/tmp/safe_harbor/{job_id}"
    os.makedirs(job_dir, exist_ok=True)
    file_path = f"{job_dir}/template.xlsx"

    content = await file.read()
    if len(content) > settings.max_file_size_mb * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large")

    with open(file_path, "wb") as f:
        f.write(content)

    orchestrator.jobs[job_id] = JobState(job_id=job_id, status="pending", scenario_type=scenario_type)
    return {"job_id": job_id, "scenario_type": scenario_type}

@app.websocket("/ws/{job_id}")
async def websocket_endpoint(websocket: WebSocket, job_id: str):
    await websocket.accept()
    
    if job_id not in orchestrator.jobs:
        await websocket.send_text(json.dumps({"error": "Job not found"}))
        await websocket.close()
        return
        
    file_path = f"/tmp/safe_harbor/{job_id}/template.xlsx"
    
    async def ws_callback(event):
        await websocket.send_text(event.model_dump_json())
        
    job = orchestrator.jobs[job_id]
    scenario_type = getattr(job, 'scenario_type', 'general') or 'general'
    try:
        await orchestrator.run_pipeline(job_id, file_path, ws_callback, scenario_type=scenario_type)
    except WebSocketDisconnect:
        logger.info(f"Client disconnected for job {job_id}")
    finally:
        pass

@app.get("/api/download/{job_id}")
async def download(job_id: str):
    if job_id not in orchestrator.jobs:
        raise HTTPException(status_code=404, detail="Job not found")
        
    job = orchestrator.jobs[job_id]
    if job.status != "complete":
        raise HTTPException(status_code=400, detail="Job not complete")
        
    return FileResponse(
        job.output_file_path, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        filename="safe_harbor_model.xlsx"
    )

@app.get("/api/audit/{job_id}")
async def get_audit(job_id: str):
    if job_id not in orchestrator.jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job = orchestrator.jobs[job_id]
    if job.status == "pending":
        raise HTTPException(status_code=400, detail="Job still pending")
        
    return job.model_dump()

@app.post("/api/sheets/template/{filename}")
async def create_template_sheet(filename: str):
    """Create a Google Sheet from a template xlsx via Sheets API."""
    file_path = f"templates/{filename}"
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Template not found")

    settings = get_settings()
    sa_path = settings.google_service_account_path
    if not os.path.exists(sa_path):
        raise HTTPException(status_code=500, detail="Google service account not configured")

    try:
        return await asyncio.to_thread(_create_sheet_from_xlsx, file_path, f"Template_{filename}", sa_path)
    except Exception as e:
        logger.error(f"Template Sheets creation failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/preview/{filename}")
async def preview_template(filename: str):
    import os
    file_path = f"templates/{filename}"
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Template not found")
    from backend.excel_io.parser import parse_template
    parsed = parse_template(file_path)
    return parsed

@app.post("/api/sheets/{job_id}")
async def create_google_sheet(job_id: str):
    """Create a Google Sheet from the output xlsx via Sheets API (no Drive upload)."""
    if job_id not in orchestrator.jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    job = orchestrator.jobs[job_id]
    if job.status != "complete" or not job.output_file_path:
        raise HTTPException(status_code=400, detail="Job not complete")

    settings = get_settings()
    sa_path = settings.google_service_account_path
    if not os.path.exists(sa_path):
        raise HTTPException(status_code=500, detail="Google service account not configured")

    try:
        return await asyncio.to_thread(
            _create_sheet_from_xlsx, job.output_file_path, f"SafeHarbor_{job_id[:8]}", sa_path, True,
            parsed_template=job.parsed_template,
            conformance_report=getattr(job, 'conformance_report', None)
        )
    except Exception as e:
        logger.error(f"Google Sheets creation failed: {e}")
        raise HTTPException(status_code=500, detail=f"Google Sheets creation failed: {str(e)}")


@app.get("/api/spreadsheet/{job_id}")
async def get_spreadsheet_data(job_id: str):
    """Return the output xlsx with computed formula values."""
    if job_id not in orchestrator.jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    job = orchestrator.jobs[job_id]
    if job.status != "complete" or not job.output_file_path:
        raise HTTPException(status_code=400, detail="Job not complete")

    import openpyxl

    # Evaluate formulas using the formulas library
    computed_values = {}
    try:
        import formulas
        xl_model = formulas.ExcelModel().loads(job.output_file_path).finish()
        solution = xl_model.calculate()
        for key, val in solution.items():
            # key is like "'Income Statement'!B6"
            import numpy as np
            v = val
            if hasattr(v, 'tolist'):
                v = v.tolist()
            if isinstance(v, (list,)):
                v = v[0] if len(v) == 1 else v
            if isinstance(v, float) and (np.isnan(v) or np.isinf(v)):
                v = None
            computed_values[str(key).upper()] = v
    except Exception as e:
        logger.warning(f"Formula evaluation failed: {e}")

    wb = openpyxl.load_workbook(job.output_file_path, data_only=False)
    sheets = []
    for ws in wb.worksheets:
        cells = {}
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                val = cell.value
                if val is not None:
                    coord = cell.coordinate
                    is_formula = isinstance(val, str) and val.startswith("=")

                    # Try to get computed value for formula cells
                    computed = None
                    if is_formula:
                        lookup = f"'{ws.title}'!{coord}".upper()
                        computed = computed_values.get(lookup)
                        if computed is None:
                            # Try without quotes
                            lookup2 = f"{ws.title}!{coord}".upper()
                            computed = computed_values.get(lookup2)

                    cells[coord] = {
                        "v": computed if is_formula else val,
                        "f": val if is_formula else None,
                        "r": row,
                        "c": col,
                    }
        sheets.append({
            "name": ws.title,
            "maxRow": max_row,
            "maxCol": max_col,
            "cells": cells,
        })
    return {"sheets": sheets}


@app.get("/api/conformance/{job_id}")
async def get_conformance(job_id: str):
    """Return the archetype conformance report for a completed job."""
    if job_id not in orchestrator.jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    job = orchestrator.jobs[job_id]
    if job.status != "complete":
        raise HTTPException(status_code=400, detail="Job not complete")
    if not job.conformance_report:
        raise HTTPException(status_code=404, detail="No conformance report available")
    return job.conformance_report


@app.get("/api/costs/{job_id}")
async def get_costs(job_id: str):
    if job_id not in orchestrator.jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job = orchestrator.jobs[job_id]
    entries = getattr(job, "cost_entries", [])
    total_cost = sum(e.estimated_cost_usd for e in entries)
    return {
        "job_id": job_id,
        "entries": [e.model_dump() for e in entries],
        "total_cost_usd": total_cost
    }
