import openpyxl
import re

class InvalidTemplateError(Exception): pass
class TemplateNotEmptyError(Exception): pass

def parse_template(file_path: str) -> dict:
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
    except Exception as e:
        raise InvalidTemplateError(f"Corrupt or unsupported Excel file: {e}")

    result = {
        "file_name": file_path.split("/")[-1],
        "sheets": [],
        "named_ranges": [],
        "inter_sheet_refs": [],
        "total_input_cells": 0
    }

    year_pattern = re.compile(r"(FY|CY)?\d{4}[EA]?")
    inter_sheet_pattern = re.compile(r"'?([^'!]+)'?!([A-Z]+\d+)")

    # Known financial acronyms that are ALL CAPS but are real data rows
    financial_acronyms = {
        'EBITDA', 'EBIT', 'EBT', 'EPS', 'ROE', 'ROA', 'ROIC', 'WACC',
        'IRR', 'MOIC', 'NPV', 'FCF', 'UFCF', 'LFCF', 'DSCR', 'SGA',
        'COGS', 'CAPEX', 'NWC', 'PP&E', 'PPE', 'D&A',
    }
    # Section header keywords
    section_keywords = {
        'activities', 'assumptions', 'summary', 'schedule', 'guide',
        'instructions', 'legend', 'disclaimer', 'notes',
    }
    # Single-word ALL CAPS section headers
    section_singles = {
        'ASSETS', 'LIABILITIES', 'EQUITY',
    }
    skip_exact = {
        'formatting guide', 'blue text', 'black text', 'green text',
        'notes', 'instructions', 'legend', 'source', 'disclaimer',
        'input', 'link to another sheet', 'formula',
    }
    skip_contains = ['color code', 'formatting', 'legend', 'instruction']

    def _is_section_header(name):
        """Detect section headers like OPERATING ACTIVITIES, TOTAL DEBT SUMMARY, etc."""
        stripped = name.strip()
        # Skip known non-data rows
        if stripped.lower() in skip_exact:
            return True
        if any(kw in stripped.lower() for kw in skip_contains):
            return True
        # Preserve known financial acronyms
        if stripped.upper() in financial_acronyms:
            return False
        # Known single-word section headers
        if stripped in section_singles:
            return True
        # ALL CAPS with 2+ words and contains a section keyword
        if stripped == stripped.upper() and len(stripped) > 5 and ' ' in stripped:
            lower = stripped.lower()
            if any(kw in lower for kw in section_keywords):
                return True
            # Generic ALL CAPS multi-word headers (like "SENIOR SECURED DEBT", "ASSETS")
            return True
        return False

    total_input = 0
    total_cells_checked = 0
    populated_input = 0

    for ws in wb.worksheets:
        sheet_data = {
            "name": ws.title,
            "headers": [],
            "input_cells": [],
            "formula_cells": [],
            "temporal_headers": []
        }

        period_headers = []
        header_row = 1
        for col in range(2, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val and year_pattern.search(str(val)):
                period_headers.append({"col": col, "val": str(val).strip()})
                if str(val).strip() not in sheet_data["temporal_headers"]:
                    sheet_data["temporal_headers"].append(str(val).strip())
        # If no periods found in row 1, try row 2
        if not period_headers:
            header_row = 2
            for col in range(2, ws.max_column + 1):
                val = ws.cell(row=2, column=col).value
                if val and year_pattern.search(str(val)):
                    period_headers.append({"col": col, "val": str(val).strip()})
                    if str(val).strip() not in sheet_data["temporal_headers"]:
                        sheet_data["temporal_headers"].append(str(val).strip())

        # If no period headers found, treat columns B+ as single-value inputs
        if not period_headers:
            for row in range(2, ws.max_row + 1):
                line_item_val = ws.cell(row=row, column=1).value
                if not line_item_val:
                    continue
                header_name = str(line_item_val).strip()
                is_section = _is_section_header(header_name)
                sheet_data["headers"].append({"row": row, "header": header_name, "is_section": is_section})
                if is_section:
                    continue
                # Scan columns B onwards for input/formula cells
                for col in range(2, min(ws.max_column + 1, 8)):  # cap at col G
                    cell = ws.cell(row=row, column=col)
                    val = cell.value
                    coord = cell.coordinate
                    total_cells_checked += 1
                    if val is None or str(val).strip() == "":
                        sheet_data["input_cells"].append({"ref": coord, "column_header": header_name, "period": "Value"})
                        total_input += 1
                    elif isinstance(val, str) and val.startswith("="):
                        sheet_data["formula_cells"].append({"ref": coord, "formula": val, "column_header": header_name})
                        matches = inter_sheet_pattern.findall(val)
                        for match in matches:
                            target_sheet, target_cell = match
                            if target_sheet != ws.title:
                                result["inter_sheet_refs"].append({
                                    "source_sheet": ws.title, "source_cell": coord,
                                    "target_sheet": target_sheet, "target_cell": target_cell
                                })
                    else:
                        populated_input += 1
                        sheet_data["input_cells"].append({"ref": coord, "column_header": header_name, "period": "Value"})
                        total_input += 1
            result["sheets"].append(sheet_data)
            continue

        data_start_row = header_row + 1
        current_section = ""  # Track which section we're in for disambiguation
        for row in range(data_start_row, ws.max_row + 1):
            line_item_val = ws.cell(row=row, column=1).value
            if not line_item_val:
                continue

            header_name = str(line_item_val).strip()
            is_section = _is_section_header(header_name)
            sheet_data["headers"].append({"row": row, "header": header_name, "is_section": is_section})

            # Track section context, skip section headers as data rows
            if is_section:
                current_section = header_name
                continue

            # Build disambiguated header: "SENIOR SECURED DEBT > Beginning Balance"
            if current_section:
                qualified_header = f"{current_section} > {header_name}"
            else:
                qualified_header = header_name

            for p in period_headers:
                cell = ws.cell(row=row, column=p["col"])
                val = cell.value
                coord = cell.coordinate

                total_cells_checked += 1
                if val is None or str(val).strip() == "":
                    sheet_data["input_cells"].append({"ref": coord, "column_header": qualified_header, "period": p["val"]})
                    total_input += 1
                elif isinstance(val, str) and val.startswith("="):
                    sheet_data["formula_cells"].append({"ref": coord, "formula": val, "column_header": header_name})

                    matches = inter_sheet_pattern.findall(val)
                    for match in matches:
                        target_sheet, target_cell = match
                        if target_sheet != ws.title:
                            result["inter_sheet_refs"].append({
                                "source_sheet": ws.title,
                                "source_cell": coord,
                                "target_sheet": target_sheet,
                                "target_cell": target_cell
                            })
                else:
                    populated_input += 1
                    sheet_data["input_cells"].append({"ref": coord, "column_header": qualified_header, "period": p["val"]})
                    total_input += 1

        result["sheets"].append(sheet_data)

    if total_input > 0 and (populated_input / total_input) > 0.15:
        raise TemplateNotEmptyError("File contains too much data in input cells. Upload an empty template.")

    result["total_input_cells"] = total_input

    # openpyxl 3.1+ uses DefinedNameDict which is directly iterable
    try:
        for name in wb.defined_names.values():
            result["named_ranges"].append({
                "name": name.name,
                "cell_range": name.attr_text
            })
    except Exception:
        pass  # No named ranges or incompatible API — not critical

    return result
