import openpyxl
from backend.models.schemas import SyntheticPayload

def write_synthetic_data(template_path: str, payload: SyntheticPayload, output_path: str) -> str:
    wb = openpyxl.load_workbook(template_path, data_only=False)
    
    for cell_value in payload.cells:
        if cell_value.sheet_name in wb.sheetnames:
            ws = wb[cell_value.sheet_name]
            cell = ws[cell_value.cell_ref]
            
            existing_val = cell.value
            if isinstance(existing_val, str) and existing_val.startswith("="):
                continue
                
            ws[cell_value.cell_ref] = cell_value.value
            
    wb.save(output_path)
    return output_path
