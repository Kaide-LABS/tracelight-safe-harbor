"""
Two-Pass Balance Sheet Constraint Projection.

Based on Coletta et al. (2023) "On the Constrained Time-Series Generation Problem"
(arXiv:2307.01717): generate first, then project onto the constraint surface.

Instead of mirroring Excel formulas in Python (which drifts), we:
1. Write the solver's output to xlsx (Pass 1 — done by writer.py)
2. Evaluate ALL template formulas using formulas.ExcelModel (the truth)
3. Read back Total Assets and Total L+E for each period
4. Compute delta and adjust a plug variable to force Assets = L + E
5. Write the corrected plug value back to the xlsx (Pass 2)
"""
import logging
import numpy as np
import openpyxl
from backend.agents.row_map import build_row_map

logger = logging.getLogger(__name__)

# The plug cell: "Other Long-Term Assets" on the Balance Sheet.
# This is an INPUT cell on the verbose LBO template (row 17).
PLUG_CANONICAL = "bs_other_noncurr"

# Fallback plug on liabilities side
FALLBACK_PLUG_CANONICAL = "bs_other_lt_liab"


def balance_bs(output_path: str, parsed_template: dict) -> str:
    """
    Two-pass BS balance: evaluate formulas, compute imbalance, write plug.

    Args:
        output_path: Path to the first-pass output xlsx (already written by writer.py)
        parsed_template: The parsed template dict from the parser

    Returns:
        The same output_path (modified in place)
    """
    if not parsed_template:
        return output_path

    rm = build_row_map(parsed_template)
    if rm["template_type"] not in ("LBO", "3-statement"):
        return output_path

    bs_name = rm["sheet_names"].get("bs")
    if not bs_name:
        return output_path

    # Row numbers for Total Assets and Total L+E
    r_total_assets = rm["row_map"].get((bs_name, "bs_total_assets"))
    r_total_le = rm["row_map"].get((bs_name, "bs_total_liab_equity"))
    r_plug = rm["row_map"].get((bs_name, PLUG_CANONICAL))
    r_fallback = rm["row_map"].get((bs_name, FALLBACK_PLUG_CANONICAL))

    if not r_total_assets or not r_total_le:
        logger.warning("BS plug: missing Total Assets or Total L+E row mapping")
        return output_path

    if not r_plug:
        logger.warning("BS plug: missing plug cell row mapping")
        return output_path

    periods = rm["periods"]
    col_letters = rm["col_letters"]

    # ── Pass 2a: Sanitize and evaluate all formulas using formulas.ExcelModel ──
    # Some templates have fake formulas like '= Formula (auto-calculated)' that
    # break the parser. Clean them before evaluation.
    import tempfile, shutil
    sanitized_path = output_path.replace(".xlsx", "_sanitized.xlsx")
    shutil.copy2(output_path, sanitized_path)

    wb_sanitize = openpyxl.load_workbook(sanitized_path, data_only=False)
    for ws in wb_sanitize.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    # Validate with formulas parser
                    try:
                        import formulas as _fml
                        _fml.Parser().ast(v, context={"name": ws.title})
                    except Exception:
                        # Not a valid formula — convert to plain string
                        cell.value = v[1:].strip()
    wb_sanitize.save(sanitized_path)

    try:
        import formulas
        xl_model = formulas.ExcelModel().loads(sanitized_path).finish()
        solution = xl_model.calculate()
    except Exception as e:
        logger.error(f"BS plug: formula evaluation failed: {e}")
        try:
            import os
            os.remove(sanitized_path)
        except:
            pass
        return output_path

    # ── Pass 2b: Read back Total Assets and Total L+E per period ──
    # Solution keys look like "'[sanitized_filename]SHEET NAME'!B20" (uppercased sheet)
    sanitized_filename = sanitized_path.split("/")[-1]

    def _read_solution(sheet_name, row, col_letter):
        """Read a computed value from the formula solution."""
        # formulas library uses: '[filename]SHEET NAME'!B20 (uppercased sheet)
        for fmt in [
            f"'[{sanitized_filename}]{sheet_name.upper()}'!{col_letter}{row}",
            f"'[{sanitized_filename}]{sheet_name}'!{col_letter}{row}",
            f"'{sheet_name}'!{col_letter}{row}",
        ]:
            val = solution.get(fmt)
            if val is None:
                continue
            # The solution values are Ranges objects — extract the actual number
            try:
                if hasattr(val, 'value'):
                    v = val.value
                    if hasattr(v, '__iter__'):
                        # It's an array like [[123.0]]
                        v = v.flat[0] if hasattr(v, 'flat') else v[0][0] if hasattr(v[0], '__iter__') else v[0]
                    v = float(v)
                else:
                    v = float(val)
                if np.isnan(v) or np.isinf(v):
                    return None
                return v
            except (TypeError, ValueError, IndexError):
                continue
        return None

    # ── Pass 2c: Compute plug adjustments ──
    wb = openpyxl.load_workbook(output_path, data_only=False)
    ws_bs = wb[bs_name]
    modified = False

    for i, col in enumerate(col_letters):
        total_assets = _read_solution(bs_name, r_total_assets, col)
        total_le = _read_solution(bs_name, r_total_le, col)

        if total_assets is None or total_le is None:
            logger.warning(f"BS plug: could not read totals for period {col}")
            continue

        delta = total_le - total_assets  # positive = L+E > Assets, need more assets

        if abs(delta) < 1.0:
            continue  # Already balanced

        # Read current plug value
        current_plug = ws_bs.cell(row=r_plug, column=i + 2).value  # col B = column 2
        if current_plug is None or (isinstance(current_plug, str) and current_plug.startswith("=")):
            current_plug = 0
        try:
            current_plug = float(current_plug)
        except (TypeError, ValueError):
            current_plug = 0

        new_plug = current_plug + delta

        if new_plug >= 0:
            ws_bs.cell(row=r_plug, column=i + 2).value = round(new_plug, 2)
            modified = True
            logger.info(f"BS plug [{col}]: Other LT Assets {current_plug:,.0f} -> {new_plug:,.0f} (delta {delta:,.0f})")
        else:
            # Can't have negative Other LT Assets — use fallback on liabilities side
            ws_bs.cell(row=r_plug, column=i + 2).value = 0
            if r_fallback:
                current_fb = ws_bs.cell(row=r_fallback, column=i + 2).value
                if current_fb is None or (isinstance(current_fb, str) and current_fb.startswith("=")):
                    current_fb = 0
                try:
                    current_fb = float(current_fb)
                except (TypeError, ValueError):
                    current_fb = 0
                # delta is positive (L+E > Assets), but we zeroed the plug, so
                # now Assets went down. Reduce liabilities to compensate.
                # New delta after zeroing plug: total_le - (total_assets - current_plug)
                new_delta = total_le - (total_assets - current_plug)
                new_fb = current_fb - new_delta
                ws_bs.cell(row=r_fallback, column=i + 2).value = round(max(0, new_fb), 2)
                logger.info(f"BS plug [{col}]: fallback — Other LT Liab {current_fb:,.0f} -> {max(0, new_fb):,.0f}")
            modified = True

    if modified:
        wb.save(output_path)
        logger.info("BS plug: wrote corrected values to xlsx")

    # Clean up sanitized temp file
    try:
        import os
        os.remove(sanitized_path)
    except:
        pass

    return output_path
