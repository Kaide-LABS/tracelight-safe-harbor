"""Verify generated templates are structurally correct."""
import openpyxl
import sys

def verify_lbo():
    wb = openpyxl.load_workbook("../templates/lbo_template.xlsx", data_only=False)
    assert len(wb.sheetnames) == 5, f"Expected 5 sheets, got {len(wb.sheetnames)}"
    assert "Income Statement" in wb.sheetnames
    assert "Debt Schedule" in wb.sheetnames
    assert "Returns Analysis" in wb.sheetnames

    # Check IS Interest Expense references DS row 16
    ws = wb["Income Statement"]
    ie_cell = ws.cell(row=9, column=2)  # Interest Expense, FY2020
    assert ie_cell.value and "Debt Schedule" in str(ie_cell.value), f"IS Interest Expense formula wrong: {ie_cell.value}"
    assert "16" in str(ie_cell.value), f"Should reference row 16, got: {ie_cell.value}"

    # Count input cells (empty non-formula cells)
    input_count = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2, values_only=False):
            for cell in row[1:]:  # skip label column
                if cell.value is None or (isinstance(cell.value, (int, float)) and cell.value == 0):
                    input_count += 1
    assert input_count > 30, f"Expected > 30 input cells, got {input_count}"
    print(f"LBO template: OK ({len(wb.sheetnames)} sheets, {input_count} input cells)")

def verify_three_statement():
    wb = openpyxl.load_workbook("../templates/three_statement_template.xlsx", data_only=False)
    assert len(wb.sheetnames) == 3, f"Expected 3 sheets, got {len(wb.sheetnames)}"
    assert "Debt Schedule" not in wb.sheetnames
    assert "Returns Analysis" not in wb.sheetnames
    print(f"3-Statement template: OK ({len(wb.sheetnames)} sheets)")

def verify_dcf():
    wb = openpyxl.load_workbook("../templates/dcf_template.xlsx", data_only=False)
    assert len(wb.sheetnames) == 4, f"Expected 4 sheets, got {len(wb.sheetnames)}"
    assert "Revenue Build" in wb.sheetnames
    assert "DCF Valuation" in wb.sheetnames
    print(f"DCF template: OK ({len(wb.sheetnames)} sheets)")

if __name__ == "__main__":
    verify_lbo()
    verify_three_statement()
    verify_dcf()
    print("\nAll templates verified successfully.")
