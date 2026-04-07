import openpyxl
from openpyxl.styles import Font, PatternFill, numbers
import os

def create_lbo():
    wb = openpyxl.Workbook()
    
    # 1. Income Statement
    ws_is = wb.active
    ws_is.title = "Income Statement"
    headers = ["Line Item", "FY2020", "FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]
    ws_is.append(headers)
    for cell in ws_is[1]:
        cell.font = Font(bold=True)
        
    line_items = [
        "Revenue", "COGS", "Gross Profit", "SG&A", "EBITDA", "D&A", "EBIT", "Interest Expense", "EBT", "Tax", "Net Income"
    ]
    
    for i, item in enumerate(line_items, 2):
        ws_is.cell(row=i, column=1, value=item)
        for col in range(2, 8):
            cell = ws_is.cell(row=i, column=col)
            cell.number_format = '#,##0'
            col_letter = openpyxl.utils.get_column_letter(col)
            
            if item == "Gross Profit":
                cell.value = f"={col_letter}2-{col_letter}3"
                cell.fill = PatternFill("solid", fgColor="E0E0E0")
            elif item == "EBITDA":
                cell.value = f"={col_letter}4-{col_letter}5"
                cell.fill = PatternFill("solid", fgColor="E0E0E0")
            elif item == "EBIT":
                cell.value = f"={col_letter}6-{col_letter}7"
                cell.fill = PatternFill("solid", fgColor="E0E0E0")
            elif item == "Interest Expense":
                cell.value = f"='Debt Schedule'!{col_letter}16"
                cell.fill = PatternFill("solid", fgColor="E0E0E0")
            elif item == "EBT":
                cell.value = f"={col_letter}8-{col_letter}9"
                cell.fill = PatternFill("solid", fgColor="E0E0E0")
            elif item == "Net Income":
                cell.value = f"={col_letter}10-{col_letter}11"
                cell.fill = PatternFill("solid", fgColor="E0E0E0")

    # 2. Balance Sheet
    ws_bs = wb.create_sheet("Balance Sheet")
    ws_bs.append(headers)
    for cell in ws_bs[1]: cell.font = Font(bold=True)
    
    bs_items = [
        "Cash", "Accounts Receivable", "Inventory", "Other Current Assets", "Total Current Assets",
        "PP&E Net", "Goodwill", "Other Non-Current Assets", "Total Assets",
        "Accounts Payable", "Accrued Expenses", "Current Portion of Debt", "Total Current Liabilities",
        "Senior Debt", "Mezzanine Debt", "Total Liabilities",
        "Common Equity", "Retained Earnings", "Total Equity", "Total Liabilities & Equity"
    ]
    
    for i, item in enumerate(bs_items, 2):
        ws_bs.cell(row=i, column=1, value=item)
        for col in range(2, 8):
            cell = ws_bs.cell(row=i, column=col)
            cell.number_format = '#,##0'
            col_letter = openpyxl.utils.get_column_letter(col)
            prev_col = openpyxl.utils.get_column_letter(col-1) if col > 2 else None
            
            if item == "Total Current Assets":
                cell.value = f"=SUM({col_letter}2:{col_letter}5)"
                cell.fill = PatternFill("solid", fgColor="E0E0E0")
            elif item == "Total Assets":
                cell.value = f"={col_letter}6+{col_letter}7+{col_letter}8+{col_letter}9"
                cell.fill = PatternFill("solid", fgColor="E0E0E0")
            elif item == "Total Current Liabilities":
                cell.value = f"=SUM({col_letter}11:{col_letter}13)"
                cell.fill = PatternFill("solid", fgColor="E0E0E0")
            elif item == "Total Liabilities":
                cell.value = f"={col_letter}14+{col_letter}15+{col_letter}16"
                cell.fill = PatternFill("solid", fgColor="E0E0E0")
            elif item == "Retained Earnings":
                if prev_col:
                    cell.value = f"={prev_col}19+'Income Statement'!{col_letter}12"
                else:
                    cell.value = f"='Income Statement'!{col_letter}12"
                cell.fill = PatternFill("solid", fgColor="E0E0E0")
            elif item == "Total Equity":
                cell.value = f"={col_letter}18+{col_letter}19"
                cell.fill = PatternFill("solid", fgColor="E0E0E0")
            elif item == "Total Liabilities & Equity":
                cell.value = f"={col_letter}17+{col_letter}20"
                cell.fill = PatternFill("solid", fgColor="E0E0E0")
                
    # 3. Cash Flow Statement
    ws_cf = wb.create_sheet("Cash Flow")
    ws_cf.append(headers)
    for cell in ws_cf[1]: cell.font = Font(bold=True)
    
    cf_items = [
        "Net Income", "D&A", "Changes in Working Capital", "Operating CF",
        "CapEx", "Investing CF",
        "Debt Drawdowns", "Debt Repayments", "Dividends", "Financing CF",
        "Net Change in Cash", "Beginning Cash", "Ending Cash"
    ]
    
    for i, item in enumerate(cf_items, 2):
        ws_cf.cell(row=i, column=1, value=item)
        for col in range(2, 8):
            cell = ws_cf.cell(row=i, column=col)
            cell.number_format = '#,##0'
            col_letter = openpyxl.utils.get_column_letter(col)
            prev_col = openpyxl.utils.get_column_letter(col-1) if col > 2 else None
            
            if item == "Net Income":
                cell.value = f"='Income Statement'!{col_letter}12"
                cell.fill = PatternFill("solid", fgColor="E0E0E0")
            elif item == "D&A":
                cell.value = f"='Income Statement'!{col_letter}7"
                cell.fill = PatternFill("solid", fgColor="E0E0E0")
            elif item == "Operating CF":
                cell.value = f"={col_letter}2+{col_letter}3+{col_letter}4"
                cell.fill = PatternFill("solid", fgColor="E0E0E0")
            elif item == "Investing CF":
                cell.value = f"=-{col_letter}6"
                cell.fill = PatternFill("solid", fgColor="E0E0E0")
            elif item == "Financing CF":
                cell.value = f"={col_letter}8-{col_letter}9-{col_letter}10"
                cell.fill = PatternFill("solid", fgColor="E0E0E0")
            elif item == "Net Change in Cash":
                cell.value = f"={col_letter}5+{col_letter}7+{col_letter}11"
                cell.fill = PatternFill("solid", fgColor="E0E0E0")
            elif item == "Beginning Cash":
                if prev_col:
                    cell.value = f"={prev_col}14"
                else:
                    cell.value = 0 # first period
                cell.fill = PatternFill("solid", fgColor="E0E0E0")
            elif item == "Ending Cash":
                cell.value = f"={col_letter}12+{col_letter}13"
                cell.fill = PatternFill("solid", fgColor="E0E0E0")

    # 4. Debt Schedule
    ws_ds = wb.create_sheet("Debt Schedule")
    ws_ds.append(headers)
    for cell in ws_ds[1]: cell.font = Font(bold=True)
    
    ds_items = [
        "Senior Debt", "Beginning Balance", "Drawdowns", "Repayments", "Ending Balance", "Interest Rate", "Interest Expense",
        "Mezzanine Debt", "Beginning Balance", "Drawdowns", "Repayments", "Ending Balance", "Interest Rate", "Interest Expense",
        "Total Interest Expense", "Total Ending Debt"
    ]
    
    for i, item in enumerate(ds_items, 2):
        ws_ds.cell(row=i, column=1, value=item)
        if item in ["Senior Debt", "Mezzanine Debt"]:
            ws_ds.cell(row=i, column=1).font = Font(bold=True)
            continue
            
        for col in range(2, 8):
            cell = ws_ds.cell(row=i, column=col)
            col_letter = openpyxl.utils.get_column_letter(col)
            prev_col = openpyxl.utils.get_column_letter(col-1) if col > 2 else None
            
            if item == "Interest Rate":
                cell.number_format = '0.0%'
            else:
                cell.number_format = '#,##0'
                
            if item == "Beginning Balance":
                if prev_col:
                    cell.value = f"={prev_col}{i+3}"
                    cell.fill = PatternFill("solid", fgColor="E0E0E0")
            elif item == "Ending Balance":
                cell.value = f"={col_letter}{i-3}+{col_letter}{i-2}-{col_letter}{i-1}"
                cell.fill = PatternFill("solid", fgColor="E0E0E0")
            elif item == "Interest Expense":
                cell.value = f"={col_letter}{i-5}*{col_letter}{i-1}"
                cell.fill = PatternFill("solid", fgColor="E0E0E0")
            elif item == "Total Interest Expense":
                cell.value = f"={col_letter}8+{col_letter}15"
                cell.fill = PatternFill("solid", fgColor="E0E0E0")
            elif item == "Total Ending Debt":
                cell.value = f"={col_letter}6+{col_letter}13"
                cell.fill = PatternFill("solid", fgColor="E0E0E0")

    # 5. Returns Analysis
    ws_ra = wb.create_sheet("Returns Analysis")
    ws_ra.append(["Metric", "Value"])
    for cell in ws_ra[1]: cell.font = Font(bold=True)
    
    ra_items = [
        ("Entry EV", ""),
        ("Exit Multiple", ""),
        ("Exit EV", "='Income Statement'!G6*B3"),
        ("Net Debt at Exit", "='Debt Schedule'!G17"),
        ("Exit Equity", "=B4-B5"),
        ("Equity Invested", ""),
        ("MOIC", "=B6/B7"),
        ("IRR", "")
    ]
    
    for i, (item, form) in enumerate(ra_items, 2):
        ws_ra.cell(row=i, column=1, value=item)
        cell = ws_ra.cell(row=i, column=2)
        if form:
            cell.value = form
            cell.fill = PatternFill("solid", fgColor="E0E0E0")
        
        if item in ["Exit Multiple", "MOIC"]:
            cell.number_format = '0.0x'
        elif item == "IRR":
            cell.number_format = '0.0%'
        else:
            cell.number_format = '#,##0'

    os.makedirs("../templates", exist_ok=True)
    wb.save("../templates/lbo_template.xlsx")


def create_three_statement():
    """3-Statement model: IS + BS + CF only. No debt schedule or returns."""
    wb = openpyxl.Workbook()
    hdrs = ["Line Item", "FY2020", "FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]
    formula_fill = PatternFill("solid", fgColor="E0E0E0")

    # Income Statement
    ws = wb.active
    ws.title = "Income Statement"
    ws.append(hdrs)
    for c in ws[1]: c.font = Font(bold=True)
    items = ["Revenue", "COGS", "Gross Profit", "SG&A", "EBITDA", "D&A", "EBIT",
             "Interest Expense", "EBT", "Tax", "Net Income"]
    for i, item in enumerate(items, 2):
        ws.cell(row=i, column=1, value=item)
        for col in range(2, 8):
            cell = ws.cell(row=i, column=col)
            cell.number_format = '#,##0'
            cl = openpyxl.utils.get_column_letter(col)
            if item == "Gross Profit":
                cell.value = f"={cl}2-{cl}3"; cell.fill = formula_fill
            elif item == "EBITDA":
                cell.value = f"={cl}4-{cl}5"; cell.fill = formula_fill
            elif item == "EBIT":
                cell.value = f"={cl}6-{cl}7"; cell.fill = formula_fill
            elif item == "EBT":
                cell.value = f"={cl}8-{cl}9"; cell.fill = formula_fill
            elif item == "Net Income":
                cell.value = f"={cl}10-{cl}11"; cell.fill = formula_fill

    # Balance Sheet (simplified — no senior/mezz split)
    ws_bs = wb.create_sheet("Balance Sheet")
    ws_bs.append(hdrs)
    for c in ws_bs[1]: c.font = Font(bold=True)
    bs = ["Cash", "Accounts Receivable", "Inventory", "Total Current Assets",
          "PP&E Net", "Total Assets",
          "Accounts Payable", "Accrued Expenses", "Debt", "Total Liabilities",
          "Common Equity", "Retained Earnings", "Total Equity", "Total Liabilities & Equity"]
    for i, item in enumerate(bs, 2):
        ws_bs.cell(row=i, column=1, value=item)
        for col in range(2, 8):
            cell = ws_bs.cell(row=i, column=col)
            cell.number_format = '#,##0'
            cl = openpyxl.utils.get_column_letter(col)
            pc = openpyxl.utils.get_column_letter(col-1) if col > 2 else None
            if item == "Total Current Assets":
                cell.value = f"=SUM({cl}2:{cl}4)"; cell.fill = formula_fill
            elif item == "Total Assets":
                cell.value = f"={cl}5+{cl}6"; cell.fill = formula_fill
            elif item == "Total Liabilities":
                cell.value = f"=SUM({cl}8:{cl}10)"; cell.fill = formula_fill
            elif item == "Retained Earnings":
                cell.value = (f"={pc}13+'Income Statement'!{cl}12" if pc
                              else f"='Income Statement'!{cl}12")
                cell.fill = formula_fill
            elif item == "Total Equity":
                cell.value = f"={cl}12+{cl}13"; cell.fill = formula_fill
            elif item == "Total Liabilities & Equity":
                cell.value = f"={cl}11+{cl}14"; cell.fill = formula_fill

    # Cash Flow
    ws_cf = wb.create_sheet("Cash Flow")
    ws_cf.append(hdrs)
    for c in ws_cf[1]: c.font = Font(bold=True)
    cf = ["Net Income", "D&A", "Changes in Working Capital", "Operating CF",
          "CapEx", "Investing CF", "Net Change in Cash", "Beginning Cash", "Ending Cash"]
    for i, item in enumerate(cf, 2):
        ws_cf.cell(row=i, column=1, value=item)
        for col in range(2, 8):
            cell = ws_cf.cell(row=i, column=col)
            cell.number_format = '#,##0'
            cl = openpyxl.utils.get_column_letter(col)
            pc = openpyxl.utils.get_column_letter(col-1) if col > 2 else None
            if item == "Net Income":
                cell.value = f"='Income Statement'!{cl}12"; cell.fill = formula_fill
            elif item == "D&A":
                cell.value = f"='Income Statement'!{cl}7"; cell.fill = formula_fill
            elif item == "Operating CF":
                cell.value = f"={cl}2+{cl}3+{cl}4"; cell.fill = formula_fill
            elif item == "Investing CF":
                cell.value = f"=-{cl}6"; cell.fill = formula_fill
            elif item == "Net Change in Cash":
                cell.value = f"={cl}5+{cl}7"; cell.fill = formula_fill
            elif item == "Beginning Cash":
                cell.value = f"={pc}10" if pc else 0; cell.fill = formula_fill
            elif item == "Ending Cash":
                cell.value = f"={cl}8+{cl}9"; cell.fill = formula_fill

    wb.save("../templates/three_statement_template.xlsx")


def create_dcf():
    """DCF model: Revenue Build + IS + FCF + DCF Valuation."""
    wb = openpyxl.Workbook()
    hdrs = ["Line Item", "FY2020", "FY2021", "FY2022", "FY2023", "FY2024", "FY2025"]
    formula_fill = PatternFill("solid", fgColor="E0E0E0")

    # Revenue Build
    ws = wb.active
    ws.title = "Revenue Build"
    ws.append(hdrs)
    for c in ws[1]: c.font = Font(bold=True)
    rev_items = ["Segment A Revenue", "Segment A Growth", "Segment B Revenue",
                 "Segment B Growth", "Total Revenue"]
    for i, item in enumerate(rev_items, 2):
        ws.cell(row=i, column=1, value=item)
        for col in range(2, 8):
            cell = ws.cell(row=i, column=col)
            cl = openpyxl.utils.get_column_letter(col)
            if "Growth" in item:
                cell.number_format = '0.0%'
            else:
                cell.number_format = '#,##0'
            if item == "Total Revenue":
                cell.value = f"={cl}2+{cl}4"; cell.fill = formula_fill

    # Income Statement
    ws_is = wb.create_sheet("Income Statement")
    ws_is.append(hdrs)
    for c in ws_is[1]: c.font = Font(bold=True)
    is_items = ["Revenue", "COGS", "Gross Profit", "SG&A", "EBITDA", "D&A",
                "EBIT", "Tax", "Net Income"]
    for i, item in enumerate(is_items, 2):
        ws_is.cell(row=i, column=1, value=item)
        for col in range(2, 8):
            cell = ws_is.cell(row=i, column=col)
            cell.number_format = '#,##0'
            cl = openpyxl.utils.get_column_letter(col)
            if item == "Revenue":
                cell.value = f"='Revenue Build'!{cl}6"; cell.fill = formula_fill
            elif item == "Gross Profit":
                cell.value = f"={cl}2-{cl}3"; cell.fill = formula_fill
            elif item == "EBITDA":
                cell.value = f"={cl}4-{cl}5"; cell.fill = formula_fill
            elif item == "EBIT":
                cell.value = f"={cl}6-{cl}7"; cell.fill = formula_fill
            elif item == "Net Income":
                cell.value = f"={cl}8-{cl}9"; cell.fill = formula_fill

    # Free Cash Flow
    ws_fcf = wb.create_sheet("Free Cash Flow")
    ws_fcf.append(hdrs)
    for c in ws_fcf[1]: c.font = Font(bold=True)
    fcf_items = ["EBITDA", "Tax", "D&A", "Changes in Working Capital", "CapEx",
                 "Unlevered FCF"]
    for i, item in enumerate(fcf_items, 2):
        ws_fcf.cell(row=i, column=1, value=item)
        for col in range(2, 8):
            cell = ws_fcf.cell(row=i, column=col)
            cell.number_format = '#,##0'
            cl = openpyxl.utils.get_column_letter(col)
            if item == "EBITDA":
                cell.value = f"='Income Statement'!{cl}6"; cell.fill = formula_fill
            elif item == "D&A":
                cell.value = f"='Income Statement'!{cl}7"; cell.fill = formula_fill
            elif item == "Unlevered FCF":
                cell.value = f"={cl}2-{cl}3+{cl}4+{cl}5-{cl}6"; cell.fill = formula_fill

    # DCF Valuation
    ws_dcf = wb.create_sheet("DCF Valuation")
    ws_dcf.append(["Metric", "Value"])
    for c in ws_dcf[1]: c.font = Font(bold=True)
    dcf_items = [
        ("WACC", ""), ("Terminal Growth Rate", ""),
        ("Terminal Value", ""), ("PV of FCFs", ""),
        ("Enterprise Value", "=B4+B5"),
    ]
    for i, (item, form) in enumerate(dcf_items, 2):
        ws_dcf.cell(row=i, column=1, value=item)
        cell = ws_dcf.cell(row=i, column=2)
        if form:
            cell.value = form; cell.fill = formula_fill
        cell.number_format = '0.0%' if "Rate" in item or item == "WACC" else '#,##0'

    wb.save("../templates/dcf_template.xlsx")


if __name__ == "__main__":
    create_lbo()
    create_three_statement()
    create_dcf()
    print("Templates generated: lbo_template.xlsx, three_statement_template.xlsx, dcf_template.xlsx")
