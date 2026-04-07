import openpyxl
import os

def create_questionnaire():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Security Assessment"
    
    headers = ["#", "Category", "Question", "Response", "Evidence"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        
    questions = [
        (1, "access_control", "Does your organization enforce multi-factor authentication for all user accounts?"),
        (2, "encryption", "Are all production databases encrypted at rest using AES-256 or equivalent?"),
        (3, "encryption", "Describe your key management practices including rotation schedules."),
        (4, "network_security", "Are there any publicly accessible endpoints other than your load balancer?"),
        (5, "network_security", "Describe your VPC segmentation and network access control strategy."),
        (6, "incident_response", "Do you have a documented incident response plan? What are your SLAs?"),
        (7, "incident_response", "Describe your breach notification procedures and timelines."),
        (8, "logging_monitoring", "Are all API calls logged? What is your log retention period?"),
        (9, "logging_monitoring", "Do you use a SIEM? Describe your monitoring and alerting capabilities."),
        (10, "data_classification", "How do you classify data? Describe your data handling tiers."),
        (11, "business_continuity", "What is your RTO and RPO for critical systems?"),
        (12, "business_continuity", "Describe your disaster recovery architecture."),
        (13, "vendor_management", "How do you assess and monitor third-party vendor risk?"),
        (14, "compliance", "Are you SOC 2 Type 2 certified? When was your last audit?"),
        (15, "compliance", "Do you comply with GDPR? How do you handle data subject requests?"),
        (16, "change_management", "Describe your change management and deployment process."),
        (17, "physical_security", "Where are your data centers located? What physical security controls exist?"),
        (18, "access_control", "Describe your RBAC model and least-privilege enforcement."),
        (19, "access_control", "How do you handle employee onboarding and offboarding access?"),
        (20, "encryption", "Is data encrypted in transit? What TLS version do you enforce?"),
        (21, "network_security", "Do you perform regular penetration testing? How often?"),
        (22, "logging_monitoring", "How do you detect unauthorized access attempts?"),
        (23, "incident_response", "Have you experienced any security breaches in the last 24 months?"),
        (24, "data_classification", "How do you handle data deletion and disposal?"),
        (25, "compliance", "Do you have cyber insurance? What is the coverage?"),
        (26, "access_control", "Do you enforce password complexity requirements? What are they?"),
        (27, "encryption", "Describe your backup encryption strategy."),
        (28, "business_continuity", "How often do you test your disaster recovery plan?"),
        (29, "vendor_management", "Do your subprocessors maintain equivalent security certifications?"),
        (30, "change_management", "How do you manage secrets and API keys in your codebase?")
    ]
    
    for row_idx, (q_id, cat, text) in enumerate(questions, 2):
        ws.cell(row=row_idx, column=1, value=q_id)
        ws.cell(row=row_idx, column=2, value=cat)
        ws.cell(row=row_idx, column=3, value=text)
        
    ws.column_dimensions['C'].width = 80
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 50
    
    os.makedirs("../tests/fixtures", exist_ok=True)
    wb.save("../tests/fixtures/sample_questionnaire.xlsx")

if __name__ == "__main__":
    create_questionnaire()
